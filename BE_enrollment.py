from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import lxml
from configparser import ConfigParser
import easygui
from easygui import choicebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import pandas as pd
import openpyxl
import os, sys


class CourseName:
    '''This class extracts from the table the name of the course and from the name of the course, extracts the name
    of the department.'''
    def __init__(self, html_table):
        self.html_table = html_table

    def pull_course_name(self):

        course_name = h2_source[table_count].text.strip()
        course_name = course_name.split()
        course_name = course_name[0] + ' ' + course_name[1]
        department = course_name.split()
        department = department[0]
        department = department.split()
        department = department[0]

        return course_name, department


class SessionName:
    row_count = 0

    def __init__(self, html_table):
        self.html_table = html_table
        # print('html table', html_table)

    def pull_session(self):
        rows = self.html_table.find_all('tr')
        # print('tr rows', rows)

        for row in rows:
            # print('row', row)
            td_row = row.find_all('td')
            # print('td_row', td_row)
            cols = [x.text.strip() for x in td_row]
            # print('cols', cols)
        if len(cols) == 2:
            for item in cols:
                if 'Session' in item:
                    session = item
                    # print('session', session)
                    return session


class TableWork:
    length = 0
    def __init__(self, html_table, course_name, department):
        self.html_table = html_table
        self.course_name = course_name
        self.deparment = department

    def extract_row(self):
        """This method extract rows the checks the length. If len is 3, then it extracts the session name. If the len
         is 33, it extracts the rest of the course information like section, instructor, size, max, etc. The output is a
         spreadsheet of each section offered in the semester."""
        cols = []
        session_row = self.html_table.find_all('td', {'colspan': '14'})
        session_row2 = self.html_table.find_all('td', {'class': 'sess2head'})
        rows = self.html_table.find_all('tr')
        for row in rows:
            if len(row) == 3:
                cols1 = row.find_all('td')
                session = [x.text.strip() for x in cols1]
                continue
            elif len(row) == 33:
                cols = row.find_all('td')
                cols = [x.text.strip() for x in cols]
                # print('cols', cols)

            else:
                continue
            cols.insert(0, self.deparment)
            cols.insert(1, self.course_name)
            cols.insert(2, session[1])
            # print(cols)
            enrollment_df.loc[TableWork.length] = cols
            TableWork.length += 1
        # print('enrollment df', enrollment_df)

class DataframeWork:

    def __init__(self, enrollment_df, semester):
        self.enrollment_df = enrollment_df
        self.semester = semester

    def sheet_integers(self):
        self.enrollment_df['Size'] = pd.to_numeric(self.enrollment_df['Size'], errors='coerce').fillna(0).astype('int')
        self.enrollment_df['Max'] = pd.to_numeric(self.enrollment_df['Max'], errors='coerce').fillna(0).astype('int')
        self.enrollment_df['Hours'] = pd.to_numeric(self.enrollment_df['Hours'], errors='coerce').fillna(0).astype('int')
        return self.enrollment_df

    def lecture_only(self):
        lecture_df = self.enrollment_df[self.enrollment_df['Type'] == 'Lecture'].reset_index()
        print(lecture_df)
        lecture_enrollment_df = lecture_df

        # lecture_enrollment_df = enrollment_df[lecture_df]

        lecture_enrollment_df.loc[lecture_enrollment_df['Room'] == 'REMOTE', 'Modality'] = 'Remote'
        lecture_enrollment_df.loc[lecture_enrollment_df['Room'] == 'ONLINE', 'Modality'] = 'Online'
        lecture_enrollment_df.loc[lecture_enrollment_df['Modality'] == 'Hybrid Course', 'Modality'] = 'Hybrid'

        rooms = ['LC 22', 'BE109', 'BE106', 'BE119', 'BE116', 'BE111', 'BE121', 'LA110', 'SS211', 'SS225', 'LA103', 'SS224',
                  'LC217','LA211', 'LA202', 'LA209', 'LA210', 'LA205', 'LA212', 'LA204', 'SS214', 'LA212', 'SS136', 'LC213',
                 'LA203', 'FA134', 'LM20*', 'FA133', 'SS136', 'BELF*', 'MAYF*', 'AHS *', 'LC134', 'SPSM*', 'WHS *', 'NHS*',
                 'STPI*', 'DOWN*', 'LA104', 'MP209', 'SS312', 'SS140', 'SS138', 'WARR*', 'AHS*', 'WHS*']
        for room in rooms:
            lecture_enrollment_df.loc[lecture_enrollment_df['Room'] == room, 'Modality'] = 'In Person'

        lecture_enrollment_df['FTES'] = lecture_enrollment_df['Size'] * (
                    ((lecture_enrollment_df['Hours'] / 18) * 17.5) / 525)
        lecture_enrollment_df['Potential FTEF'] = lecture_enrollment_df['Max'] * (((lecture_enrollment_df['Hours'] / 18) * 17.5) / 525)
        lecture_enrollment_df['FTEF'] = (lecture_enrollment_df['Hours'] / 18)/ 15
        lecture_enrollment_df['Efficiency'] = lecture_enrollment_df['FTES'] / lecture_enrollment_df['FTEF']
        lecture_enrollment_df.reset_index()
        print('lecture df', lecture_enrollment_df)
        for i in range(len(lecture_enrollment_df)):
            if 'Regular' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i, 'Session'] = '18'
            elif 'Fifteen Week A ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15A'
            elif 'Fifteen Week A1 Monday Session' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15A'
            elif 'Fifteen Week A2 Monday Session' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15A'
            elif 'Fifteen Week A3 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15A'
            elif 'Fifteen Week A6 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15A'
            elif 'Fifteen Week B M-F Session' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15B'
            elif 'Fifteen Week B6 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15B'
            elif 'Fifteen Week B7 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15B'
            elif 'Fifteen Week B1 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15B'
            elif 'Fifteen Week B2 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15B'
            elif 'Fifteen Week B16 ' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '15B'


            elif 'Nine Week A M-F Session (1/9/2023 - 3/10/2023)' \
                    in lecture_enrollment_df.loc[i, 'Session']:
                    lecture_enrollment_df.loc[i, 'Session'] = '9A'
            elif 'Nine Week A7 T,Th Session' \
                    in lecture_enrollment_df.loc[i, 'Session']:
                    lecture_enrollment_df.loc[i, 'Session'] = '9A'
            elif 'Nine Week A4 Thursday Session (1/12/2023 - 3/9/2023)' \
                    in lecture_enrollment_df.loc[i, 'Session']:
                    lecture_enrollment_df.loc[i, 'Session'] = '9A'

            elif 'Nine Week B M-F Session' \
                    in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '9B'
            elif 'Nine Week B2 Tuesday Session' \
                    in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '9B'
            elif 'Nine Week B6 M,W, Session' \
                    in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '9B'
            elif 'Nine Week B7 T, Th Session' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '9B'
            elif 'Sixteen' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i, 'Session'] = '16'
            elif 'Twelve' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '12'
            elif 'Six' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = '6'
            elif 'Enrollment' in lecture_enrollment_df.loc[i, 'Session']:
                lecture_enrollment_df.loc[i,'Session'] = 'Open'

        print('clean lecture enrollment', lecture_enrollment_df)
        if 'Fall' in self.semester:
            print(lecture_enrollment_df)
            lecture_enrollment_df.to_csv('C:/Users/fmixson/Desktop/Dashboard_files/Fall_Division_Enrollment.csv')
            lecture_enrollment_df.to_csv('Fall_Division_Enrollment.csv')

        else:
            lecture_enrollment_df.to_csv('C:/Users/fmixson/Desktop/Dashboard_files/Business_Ed_Spring_Division_Enrollment.csv')
            lecture_enrollment_df.to_csv('Business_Ed_Spring_Division_Enrollment.csv')
            lecture_enrollment_df.to_excel('Business_Ed_Spring_Division_Enrollment.xlsx')
        # grp = lecture_enrollment_df.get_group('Room')
        # divModalities = lecture_enrollment_df.groupby(['Room'])['Size', 'Max'].agg(['count', 'sum'])
        # divModalities = divModalities.to_excel('Division_Modalities.xlsx')

        return


class GroupDepartments:
    headers = ['Dept', 'Course', 'Session', 'Class', 'Start', 'End', 'Days', 'Room', 'Size', 'Max', 'Wait', 'Cap',
               'Seats', 'WaitAv', 'Status', 'Instructor', 'Type', 'Hours', 'Books', 'Modality']
    df = pd.DataFrame()
    pd.set_option('display.max_columns', None)

    def __init__(self, department, df):
        self.department = department
        self.df = df

    def group_by_departments(self):
        # workbook = openpyxl.Workbook()
        # ws1 = workbook.active
        # ws2 = workbook.create_sheet('TWO')
        grp = self.df.get_group(department)
        # ws2['A1'] = 'ID'
        # grp.to_excel(department + '/' + department + '.xlsx')
        # grp_courses = grp.groupby(['Course'])['Size', 'Max'].agg(['count', 'sum'])
        # mod_grp_courses = grp.groupby(['Room'])['Size', 'Max'].agg(['count', 'sum', 'mean'])
        # grp_courses.to_excel(department + '/' + department + 'ttl.xlsx')
        # mod_grp_courses.to_excel(department + '/' + department + '_modalities.xlsx')

# msg = 'What semester, first or second?'
# title = 'Enrollment Data'
# choices = ['1', '2']
# choice = choicebox(msg, title, choices)
s = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)
driver.get('https://secure.cerritos.edu/schedule/')
# /html/body/form/p[1]/label[1]/input
# /html/body/form/p[1]/label[2]/input
# print(choice)
#trying something new
page_loading = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'divisions')))
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'lxml')
tag = soup.p
labelTag = tag.find_all('label')


semesters = []
for label in labelTag:
    print(label)
    semesters.append(label.text[:-1])
print('semester', semesters)
# msg = 'For what semester do you want the enrollment?'
# title = 'Current Enrollment Data'
# choices = ['1', '2']
# user_choice = choicebox(msg, title, semesters)
# if user_choice == semesters[0]:
#      choice = 1
# else:
#      choice = 2
global semester_glob
for semester in semesters:
    semester_glob = semester
    if "Spring" in semester:
        driver.get('https://secure.cerritos.edu/schedule/')
        page_loading = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'divisions')))
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'lxml')
        page_loading = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'divisions')))
        # semester = driver.find_element(By.XPATH, 'html/body/form/p[1]/label[2]/input').click()
        page_loading = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'divisions')))
        semester = driver.find_element(By.XPATH, 'html/body/form/p[1]/label[1]/input').click()
        check_all = driver.find_element(By.XPATH, '/html/body/form/table[1]/tbody/tr[1]/td[1]/label/input').click()
        check_LA = driver.find_element(By.XPATH, 'html/body/form/table[6]/tbody/tr[2]/td[1]/label/input').click()

        # semester = driver.find_element(By.XPATH,'html/body/form/p[1]/label[2]/input').click()
    else:
        semester = driver.find_element(By.XPATH, 'html/body/form/p[1]/label[2]/input').click()
        semester = driver.find_element(By.XPATH, 'html/body/form/p[1]/label[2]/input').click()
        check_all = driver.find_element(By.XPATH, '/html/body/form/table[1]/tbody/tr[1]/td[1]/label/input').click()
        check_LA = driver.find_element(By.XPATH, '/html/body/form/table[6]/tbody/tr[2]/td[1]/label/input').click()
    # if choice == 2:
    #     check_LA = driver.find_element(By.XPATH,'html/body/form/table[6]/tbody/tr[5]/td[2]/label/input').click()
    # else:
    #     check_LA = driver.find_element(By.XPATH, '/html/body/form/table[6]/tbody/tr[5]/td[2]/label/input').click()
        # / html / body / form / table[6] / tbody / tr[1] / td[3] / label / input
    click_View = driver.find_element(By.XPATH, '/html/body/form/p[4]/input').click()
    # /html/body/form/table[6]/tbody/tr[5]/td[2]/label/input
    page_loading = WebDriverWait(driver, 10)
        # .until(EC.presence_of_element_located((By.NAME, 'descs')))
    headers = ['Dept', 'Course', 'Session', 'Class', 'Start', 'End', 'Days', 'Room', 'Size', 'Max', 'Wait', 'Cap', 'Seats',
               'WaitAv', 'Status', 'Instructor', 'Type', 'Hours', 'Books', 'Modality']
    pd.set_option('display.max_columns', None)
    enrollment_df = pd.DataFrame(columns=headers)
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'lxml')
    table_count = 0
    h2_source = soup.find_all('h2')
    session_source = soup.find_all('tr', {'class': 'sess1head', 'colspan': '14'})
    'Each table consists of a course section'
    tables = soup.find_all(['table', {'cellspacing': '0', 'class': 'class'}])

    for table in tables:
        'The for loop extracts from each table information about the course'
        c = CourseName(html_table=h2_source)
        course_name, department = c.pull_course_name()
        t = TableWork(html_table=table, course_name=course_name, department=department)
        t.extract_row()
        table_count += 1
    d = DataframeWork(enrollment_df=enrollment_df, semester=semester_glob)
    d.sheet_integers()
    d.lecture_only()