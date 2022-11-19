import pandas as pd
from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet('MySheet')
ws2 = wb.create_sheet('Mysheet2', 0)

enrollmentDF = pd.read_excel('Division_Enrollment.xlsx')
# print(enrollmentDF)

deptSections = enrollmentDF.groupby('Dept').Course.count()
print(deptSections)
deptEnrollment = enrollmentDF.groupby('Dept').Size.sum()
print(deptEnrollment)
deptCaps = enrollmentDF.groupby('Dept').Max.sum()
print(deptCaps)

aggData = enrollmentDF.groupby('Dept').agg({'Course':'count', 'Size':'sum', 'Max':'sum'})
print(aggData)

modalities = enrollmentDF.groupby('Room').count()
modalitiesSize = enrollmentDF.groupby('Room').Size.sum()
modalitiesMax = enrollmentDF.groupby('Room').Max.sum()
print(modalities)
print(modalitiesSize)
print(modalitiesMax)